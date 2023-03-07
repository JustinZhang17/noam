from fastapi import FastAPI
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import uvicorn
import ast
import random
import hashlib

import os.path
from dotenv import dotenv_values
import boto3

wb = Workbook()
ws = wb.active

filePath = "../assets/word_bank.txt"
app = FastAPI()

apiKeys = ["7368e427085899ce0061d35c3783a475025e99e978e1cfd6f2074d8ebb3ae65e8b3a74f2450fed439679d93bc849fe7a67514b6cd6cf282988359a7f72e83f83"]


def proStr(arr=[]):
    s = ""
    for pro in arr:
        s += "\\ " + pro + " \\" + " or "
    return s[:-4]


def posStr(arr=[]):
    s = ""
    for pos in arr:
        s += pos + ", "
    return s[:-2]


# 6yW6IJMyJhhwBUZTbev7SHG6zJpUirgZ TODO: Change this and the api hash
@app.get("/wordlist")
async def getWordList(apiKey: str = "", size: int = 200):
    if (size < 1 or size > 600):
        return {"Status_Code": 404}

    if (hashlib.sha512(apiKey.encode()).hexdigest() not in apiKeys):
        return {"Status_Code": 404}

    listName = random.getrandbits(64)

    lines = open("../assets/word_bank.txt", "r",
                 encoding='UTF8').read().splitlines()
    random.shuffle(lines)

    ws["A1"] = 'Word'
    ws["B1"] = 'Pronounciation'
    ws["C1"] = 'Part Of Speech'
    ws["D1"] = 'Definition'
    ws["E1"] = 'Sentence'
    ws["F1"] = 'Etymology'

    ws["A1"].font = Font(size=13, bold=True, name='Times New Roman')
    ws["B1"].font = Font(size=13, bold=True, name='Times New Roman')
    ws["C1"].font = Font(size=13, bold=True, name='Times New Roman')
    ws["D1"].font = Font(size=13, bold=True, name='Times New Roman')
    ws["E1"].font = Font(size=13, bold=True, name='Times New Roman')
    ws["F1"].font = Font(size=13, bold=True, name='Times New Roman')

    for i in range(size):
        term = ast.literal_eval(lines[i])
        ws["A" + str(i+2)] = term['name']
        ws["B" + str(i+2)] = proStr(term['pronounciation'])
        ws["C" + str(i+2)] = posStr(term['partOfSpeech'])
        ws["D" + str(i+2)] = term['definition']
        ws["E" + str(i+2)] = term['sentences'][0]
        ws["F" + str(i+2)] = term['etymology']

        ws["A" + str(i+2)].font = Font(name='Times New Roman')
        ws["B" + str(i+2)].font = Font(name='Times New Roman')
        ws["C" + str(i+2)].font = Font(name='Times New Roman')
        ws["D" + str(i+2)].font = Font(name='Times New Roman')
        ws["E" + str(i+2)].font = Font(name='Times New Roman')
        ws["F" + str(i+2)].font = Font(name='Times New Roman')

    wb.save(str(listName) + ".xlsx")
    return FileResponse(path=(str(listName) + ".xlsx"), filename="Noam-Wordlist-" + str(size) + ".xlsx")

if __name__ == "__main__":
    # Environment Check
    environ = ".env"
    if (os.path.isfile(".env.local")):
        environ = ".env.local"

    cred = dotenv_values(environ)

    session = boto3.Session(
        aws_access_key_id=cred.get("AWS_ACCESS_KEY_ID"),
        aws_secret_access_key=cred.get("AWS_SECRET_ACCESS_KEY"),
        region_name=cred.get("AWS_DEFAULT_REGION"),
    )

    dynaDB = session.resource('dynamodb')
    table = dynaDB.Table('noam-wordlist')

    lines = open("../assets/word_bank.txt", "r",
                 encoding='UTF8').read().splitlines()

# TODO: Remove DynamoDB table from AWS & Codebase
# TODO: Remove .env files
# TODO: Remove DynamoDB Access from account (Access Policies)

    # with table.batch_writer() as batch:
    #     for i in range(0, 16750, 25):
    #         for j in range(0, 25):
    #             print(i+j)
    #             batch.put_item(
    #                 Item=ast.literal_eval(lines[i + j])
    #             )
    # for i in range(16750, 16773):
    #     print(i)
    #     table.put_item(
    #         Item=ast.literal_eval(lines[i])
    #     )

    # uvicorn.run("main:app", reload=True)

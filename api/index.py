import os
from os.path import join
from fastapi import BackgroundTasks, FastAPI, Response, status
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font
import ast
import random
import hashlib

file_type = ".xlsx"
app = FastAPI()

apiKeys = ["94dbb126ff4e9700d201333b897b027505dfd2874f873b947f7e374592e3d96ddb369d9014908257decdf965a327582da4774cc5b687c386dd1a9785fafa3605"]


def pro_str(arr=[]):
    s = ""
    for pro in arr:
        s += "\\ " + pro + " \\" + " or "
    return s[:-4]


def pos_str(arr=[]):
    s = ""
    for pos in arr:
        s += pos + ", "
    return s[:-2]


@app.get("/wordlist")
async def get_wordlist(apiKey: str = "", size: int = 200, background_tasks: BackgroundTasks = BackgroundTasks, response: Response = Response):

    hashed: str = hashlib.sha512(apiKey.encode()).hexdigest()

    print("Hashed: " + hashed)

    if (size < 1 or size > 600 or hashed not in apiKeys):
        response.status_code = status.HTTP_401_UNAUTHORIZED
        return {"error": "Invalid API Key or Size"}

    list_name: int = random.getrandbits(64)

    lines: list = open(join('data', 'word_bank.txt'), "r",
                       encoding='UTF8').read().splitlines()

    random.shuffle(lines)

    wb = Workbook()
    ws = wb.active

    font_family: str = "Times New Roman"

    set_sheet_style(ws, font_family)

    for i in range(size):
        term = ast.literal_eval(lines[i])
        ws["A" + str(i+2)] = term['name']
        ws["B" + str(i+2)] = pro_str(term['pronunciation'])
        ws["C" + str(i+2)] = pos_str(term['partOfSpeech'])
        ws["D" + str(i+2)] = term['definition']
        ws["E" + str(i+2)] = term['sentences'][0]
        ws["F" + str(i+2)] = term['etymology']

        ws["A" + str(i+2)].font = Font(name=font_family)
        ws["B" + str(i+2)].font = Font(name=font_family)
        ws["C" + str(i+2)].font = Font(name=font_family)
        ws["D" + str(i+2)].font = Font(name=font_family)
        ws["E" + str(i+2)].font = Font(name=font_family)
        ws["F" + str(i+2)].font = Font(name=font_family)

    wb.save('/tmp/' + str(list_name) + file_type)

    # Runs after the response is sent
    background_tasks.add_task(delete_file, list_name)
    background_tasks.add_task(wb.close)

    print("File Created: " + str(list_name) + file_type)
    return FileResponse(path=('/tmp/' + str(list_name) + file_type), filename="Noam-Wordlist-" + str(size) + file_type, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def set_sheet_style(ws, font_family: str):
    ws["A1"] = 'Word'
    ws["B1"] = 'Pronunciation'
    ws["C1"] = 'Part Of Speech'
    ws["D1"] = 'Definition'
    ws["E1"] = 'Sentence'
    ws["F1"] = 'Etymology'

    ws["A1"].font = Font(size=13, bold=True, name=font_family)
    ws["B1"].font = Font(size=13, bold=True, name=font_family)
    ws["C1"].font = Font(size=13, bold=True, name=font_family)
    ws["D1"].font = Font(size=13, bold=True, name=font_family)
    ws["E1"].font = Font(size=13, bold=True, name=font_family)
    ws["F1"].font = Font(size=13, bold=True, name=font_family)


def delete_file(list_name: int):
    os.remove('/tmp/' + str(list_name) + file_type)
